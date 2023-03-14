# Import tkinter and other required libraries
import tkinter as tk
from PIL import ImageTk, Image
from tkinter import filedialog
from tkinter import PhotoImage
from tkinter import messagebox
import numpy as np
import pickle
from joblib import load
from openpyxl import Workbook, load_workbook
import os

# Create the root window
root = tk.Tk()

# Set the size of the root window
root.geometry('800x600')
root.title('Just Taxi Prediction App')

# Disable resizing the GUI
root.resizable(False, False)

# Set the icon of the root window
root.iconbitmap('icons/taxi_logo.ico')

# Load Pipeline
with open('Saved_Models/pca_pipeline.pkl', 'rb') as p:
    preprocess = pickle.load(p)

# Load Final Model
with open('Saved_Models/final_model.pkl', 'rb') as m:
    model = pickle.load(m)

def rt_mode():
    def submission():
        entries2 = []
        # Allow prompt to be properly destroyed and reused
        nonlocal prompt
        try:
            prompt.destroy()
        except NameError:
            pass

        # Empty label to leave whitespace
        prompt = tk.Label(form, text=f"", font=("bold", 15), fg = 'red')
        prompt.grid(row = 1, column = 0, columnspan = 2, pady = 8, padx = root.winfo_width() / 8 )

        # Check if all entries are filled
        for i in range(len(entries)):
            if entries[i].get() == '':
                prompt = tk.Label(form, text=f"Please Enter a Value for Feature \n - {features[i]}", font=("bold", 15), fg = 'red')
                prompt.grid(row = 1, column = 0, columnspan = 2, pady = 8, padx = root.winfo_width() / 8 )

            try:
                float(entries[i].get())
                isnum = True
            except:
                try: 
                    int(entries[i].get())
                    isnum = True
                except:
                    isnum = False

            if not isnum:
                prompt = tk.Label(form, text=f"Please Enter a Numerical Value for Feature \n - {features[i]}", font=("bold", 12), fg = 'red')
                prompt.grid(row = 1, column = 0, columnspan = 2, pady = 3, padx = root.winfo_width() / 8 )

            else:
                entries2.append(float(entries[i].get()))

        # If all features are entered and valid, then make prediction
        if len(entries2) == 33:
            # Preprocess the entered features
            X_train = preprocess.transform([entries2])
            # Make prediction
            y_pred = model.predict(X_train)
            # Possible prediction types
            p_type = ['Normal', 'Dangerous']

            # Show prediction in message box
            messagebox.showinfo('Prediction', f'Prediction: {p_type[int(y_pred[0])]}')


    for widget in root.winfo_children():
        widget.destroy()

    # Create Frame
    frame = tk.Frame(root)
    frame.pack(fill = 'both', expand = True)

    # Create Canvas
    canvas = tk.Canvas(frame, highlightthickness=0)
    canvas.pack(side = 'left', fill = 'both', expand = True)

    # Add scrollbar to the canvas
    scrollbar1 = tk.Scrollbar(frame, command = canvas.yview, orient='vertical')
    scrollbar1.pack(side="right", fill="y")

    # Configure the canvas
    canvas.configure(yscrollcommand = scrollbar1.set)
    canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion = canvas.bbox('all')))

    # Create another frame inside the canvas
    form = tk.Frame(canvas)

    # Add that new frame to a window in the canvas
    canvas.create_window((0, 0), window = form, anchor = 'nw')

    tk.Label(form, text="Please Enter Featues for Prediction", font=("bold", 18)).grid(row = 0, column = 0, columnspan = 2, pady = 10, padx = root.winfo_width() / 4.5)
    
    prompt = tk.Label(form, text="")
    prompt.grid(row = 1, column = 0, columnspan = 2, pady = 5, padx = root.winfo_width() / 8 )

    entries = []
    features = ['Mean Accuracy', 'Max Accuracy', 'Mean Bearing', 'Median Bearing', 
            'Mean Acceleration_x', 'Median Acceleration_x', 'Min Acceleration_x', 'Max Acceleration_x', 'Mean Acceleration_y', 'Median Acceleration_y', 'Min Acceleration_y', 'Max Acceleration_y',
            'Median Acceleration_z', 'Min Acceleration_z', 'Max Acceleration_z', 'Mean Gyroscope_x', 'Median Gyroscope_x', 'Min Gyroscope_x', 'Max Gyroscope_x',
            'Mean Gyroscope_y', 'Median Gyroscope_y', 'Min Gyroscope_y', 'Max Gyroscope_y', 'Mean Gyroscope_z', 'Median Gyroscope_z', 'Min Gyroscope_z', 'Max Gyroscope_z',
            'Mean Second', 'Median Second', 'Max Second', 'Mean Speed', 'Median Speed', 'Max Speed']
    for index in range(2, len(features) + 2):
        # Create spacing between features
        if (index - 1) % 4 == 0:
            text = tk.Label(form, text = features[index - 2], font = 15).grid(row = index, column = 0, padx = 8, pady = (10, 50), sticky = 'e')
            entry = tk.Entry(form)
            entries.append(entry)
            entry.grid(row = index, column = 1, pady = (8, 40), padx = 5, sticky = 'w')
        else:
            text = tk.Label(form, text = features[index - 2], font = 15).grid(row = index, column = 0, padx = 8, pady = 10, sticky = 'e')
            entry = tk.Entry(form)
            entries.append(entry)
            entry.grid(row = index, column = 1, pady = 8, padx = 5, sticky = 'w')

    submit = tk.Button(form, text = 'Submit', font = 15, padx = 20, pady = 20, bg = "black", fg = 'white', command = submission)
    submit.grid(row = len(features) + 2, column = 1, pady = 20)

    home = tk.Button(form, text = 'Home', font = 15, padx = 20, pady = 20, bg = "black", fg = 'white', command = main)
    home.grid(row = len(features) + 2, column = 0)

    form.grid_columnconfigure(0, weight=1)
    form.grid_columnconfigure(1, weight=1)


def excel_mode():
    root.rowconfigure(0, weight = 1)
    root.rowconfigure(1 , weight = 1)
    root.rowconfigure(2, weight = 1)
    root.rowconfigure(3, weight = 4)
    root.rowconfigure(4, weight = 1)
    root.columnconfigure(2, weight = 1)
    root.columnconfigure(3, weight = 1)

    root.configure(bg = '#82b7dc')

    def load_excel():
        # Enable prompt to be destroyed and recreated properly
        nonlocal prompt
        records = {}
        excel_data = filedialog.askopenfilename(initialdir = "/", title = "Select Excel File", filetypes = (("Excel Files (.xlsx)", "*.xlsx"), ("Old Excel FIles (.xls)", "*.xls*")))
        choose_excel = load_workbook(excel_data, data_only = True)
        sheet = choose_excel.active

        # Check if all features are filled
        for i in range(2, sheet.max_row + 1):
            records[i - 1] = []
            if len(sheet[i]) != 34:
                try:
                    prompt.destroy()
                except NameError:
                    pass

                prompt = tk.Label(root, text = "Please ensure all 34 features are filled", font = 15, bg = "#82b7dc", fg = 'red')
                prompt.grid(row = 2, column = 0, sticky = 'new', columnspan = 4)
                break

            # Check if all features are of integers or floats
            for r in range(1, len(sheet[i])):
                if type(sheet[i][r].value) is not int and type(sheet[i][r].value) is not float:
                    try:
                        prompt.destroy()
                    except NameError:
                        pass

                    prompt = tk.Label(root, text = "Please ensure all features are of integers or floats", font = 15, bg = "#82b7dc", fg = 'red')
                    prompt.grid(row = 2, column = 0, sticky = 'new', columnspan = 4)
                    break
                
                # else:
                records[i - 1].append(sheet[i][r].value)
 
        # Predict results
        result = []    
        for index in records:
            X_train = preprocess.transform([records[index]])
            y_pred = model.predict(X_train)
            status = ['Safe', 'Unsafe']
            result.append(status[int(y_pred[0])])
        
        try:
            prompt.destroy()
        except NameError:
            pass

        ttt = []
        for i in range(len(result)):
            ttt.append((str(i+1), result[i]))

        try:
            prompt.destroy()
        except NameError:
            pass
        prompt = tk.Label(root, text = "Results are shown here:", font = 15, bg = "#82b7dc", fg = 'green')
        prompt.grid(row = 2, column = 0, sticky = 'new', columnspan = 4)

        frame = tk.Frame(root, background='#82b7dc')
        frame.grid(row=3, column=0, columnspan = 4, sticky='news', padx = 140)

        # Create Canvas
        canvas = tk.Canvas(frame, highlightthickness=0, width=485, bg='#82b7dc')
        canvas.grid(row=0, column=0, columnspan = 3, sticky='news')

        # Add scrollbar to the canvas
        scrollbar1 = tk.Scrollbar(frame, command = canvas.yview, orient='vertical')
        scrollbar1.grid(row=0, column=4, sticky='ns')

        # Configure the canvas
        canvas.configure(yscrollcommand = scrollbar1.set)
        canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion = canvas.bbox('all')))

        # Create another frame inside the canvas
        form = tk.Frame(canvas)

        canvas.create_window((0, 0), window = form, anchor = 'nw')

        # Create the table
        ele = tk.Entry(form, font = ('Arial', 16), justify=tk.CENTER)
        ele.grid(row = 0, column = 0, sticky = 'ne')
        ele.insert(tk.END, 'Booking ID')
        ele.config(state= "disabled", disabledforeground='blue')

        ele = tk.Entry(form, font = ('Arial', 16), justify=tk.CENTER)
        ele.grid(row = 0, column = 1, sticky = 'nw')
        ele.insert(tk.END, 'Status')
        ele.config(state= "disabled", disabledforeground='blue')

        for i in range(len(ttt)):
            for j in range(len(ttt[0])):
                if ttt[i][j] == 'Safe':
                    ele = tk.Entry(form, font = ('Arial', 16), justify=tk.CENTER)
                    ele.insert(tk.END, ttt[i][j])
                    ele.config(state= "disabled", disabledforeground='green')
                else:
                    ele = tk.Entry(form, font = ('Arial', 16), justify=tk.CENTER)
                    ele.insert(tk.END, ttt[i][j])
                    ele.config(state= "disabled", disabledforeground='red')
                # ele = tk.Entry(res_grid, fg = 'blue', font = ('Arial', 16))
                if j == 0:
                    ele = tk.Entry(form, font = ('Arial', 16), justify=tk.CENTER)
                    ele.grid(row = i+1, column = j, sticky = 'ne')
                    ele.insert(tk.END, ttt[i][j])
                    ele.config(state= "disabled", disabledforeground='blue')
                elif j == 1:
                    ele.grid(row = i+1, column = j, sticky = 'nw')

                form.columnconfigure(0, weight = 1)
                form.columnconfigure(1, weight = 1)

        next_col = sheet.max_column + 1
        cell = sheet.cell(row = 1, column = next_col)
        cell.value = 'Status'
        
        for i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row = i, column = next_col)
            cell.value = result[i - 2]  

        # Save the excel file
        response = messagebox.askyesno('Save Excel File', 'Do you want to save the updated excel file?')  
        if response == 1:
            files = (("Excel Files (.xlsx)", "*.xlsx"), ("Old Excel FIles (.xls)", "*.xls*"))
            file_path = filedialog.asksaveasfilename(initialdir = "/", title = "Select Excel File", filetypes = files, defaultextension = files)
            choose_excel.save(file_path)   

    
    def download_excel():
        wb = Workbook()
        ws = wb.active

        headers = ['BookingID', 'Mean Accuracy', 'Max Accuracy', 'Mean Bearing', 'Median Bearing', 
            'Mean Acceleration_x', 'Median Acceleration_x', 'Min Acceleration_x', 'Max Acceleration_x', 'Mean Acceleration_y', 'Median Acceleration_y', 'Min Acceleration_y', 'Max Acceleration_y',
            'Median Acceleration_z', 'Min Acceleration_z', 'Max Acceleration_z', 'Mean Gyroscope_x', 'Median Gyroscope_x', 'Min Gyroscope_x', 'Max Gyroscope_x',
            'Mean Gyroscope_y', 'Median Gyroscope_y', 'Min Gyroscope_y', 'Max Gyroscope_y', 'Mean Gyroscope_z', 'Median Gyroscope_z', 'Min Gyroscope_z', 'Max Gyroscope_z',
            'Mean Second', 'Median Second', 'Max Second', 'Mean Speed', 'Median Speed', 'Max Speed']

        # Write headers to the Excel sheet
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header

        ws['A2'] = 1
        ws['A3'] = 2

        files = (("Excel Files (.xlsx)", "*.xlsx"), ("Old Excel FIles (.xls)", "*.xls*"))
        file_path = filedialog.asksaveasfilename(initialdir = "/", title = "Select Excel File", filetypes = files, defaultextension = files)
        wb.save(file_path)

        # Save the excel file
        if file_path != '':
            nonlocal prompt
            try:
                prompt.destroy()
            except NameError:
                pass

            prompt = tk.Label(root, text = "Template Saved", font = 15, bg = "#82b7dc", fg = 'green')
            prompt.grid(row = 2, column = 0, sticky = 'new', columnspan = 4)

        

    for widget in root.winfo_children():
        widget.destroy()

    label = tk.Label(root, text = "Batch Prediction Mode", font = 15, pady = 20, bg = "#82b7dc")
    label.grid(row = 0, column = 0, sticky = 'sew', columnspan = 4)

    instructions = tk.Label(root, text = "Please use template provided for prediction.", font = 15, bg = "#82b7dc")
    instructions.grid(row = 1, column = 0, sticky = 'new', columnspan = 4)

    excel_template = tk.Button(root, text = 'Download Excel Template', font = 15, padx = 15, pady = 20, bg = "black", fg = 'white', command = download_excel)
    excel_template.grid(row = 4, column = 1)

    prompt = tk.Label(root, text = "", font = 15, bg = "#82b7dc", fg = 'green')
    prompt.grid(row = 2, column = 0, sticky = 'new', columnspan = 4)

    choose_excel = tk.Button(root, text = 'Select Excel File', font = 15, padx = 15, pady = 20, bg = "black", fg = 'white', command = load_excel)
    choose_excel.grid(row = 4, column = 2)

    home = tk.Button(root, text = 'Home', font = 15, padx = 80, pady = 20, bg = "black", fg = 'white', command = main)
    home.grid(row = 4, column = 0)

def main():
    for widget in root.winfo_children():
        widget.destroy()
    
    for i in range(5):
        root.rowconfigure(i, weight = 0)
        root.columnconfigure(i, weight= 0)

    root.rowconfigure(0, weight = 1)
    root.rowconfigure(1, weight = 2)
    root.rowconfigure(2, weight = 2)
    root.columnconfigure(0, weight=1)
    root.columnconfigure(1, weight=1)
    root['bg'] = '#fcfcd4'

    label = tk.Label(root, text="Trip Safety Prediction", font=("Arial", 25), bg = '#fcfcd4')
    label.grid(row = 0, column = 0, sticky = 'ews', columnspan = 2, pady = 10)

    label = tk.Label(root, text="This is a prediction app that predicts whether a trip is safe or unsafe \nbased on your input parameters", font=("Arial", 15), bg = '#fcfcd4')
    label.grid(row = 1, column = 0, sticky = 'new', columnspan = 2)

    choose_mode1 = tk.Button(root, text='Real-Time Mode', font = 15,
                            padx=35, pady=20,   # Box padding
                            fg="white", bg="black", command=rt_mode)
    choose_mode1.grid(row = 2, column = 0)

    choose_mode2 = tk.Button(root, text='Batch Mode', font = 15,
                            padx=35, pady=20,
                            fg="white", bg="black", command=excel_mode)
    choose_mode2.grid(row = 2, column = 1)

main()
root.mainloop()